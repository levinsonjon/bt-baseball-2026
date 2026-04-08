"""
pitcherlist.py — Load projected stats from a PitcherList Excel export.

Expected file: data/pitcherlist_projections.xlsx

To update projections, replace that file with a newer PitcherList export in the
same format. The loader reads the "Pitchers" and "Hitters" sheets and expects
the column headers listed in HITTER_COLUMNS and PITCHER_COLUMNS below.

The file is gitignored. Keep a copy outside the repo if needed.
"""

from __future__ import annotations
from pathlib import Path
from typing import Optional

import openpyxl

from players import Player

DATA_DIR = Path(__file__).parent / "data"
DEFAULT_FILE = DATA_DIR / "pitcherlist_projections.xlsx"

# Columns we read from each sheet (must match the Excel header row exactly).
# Any extra columns in the file are ignored.
HITTER_COLUMNS = {
    "Name": "name",
    "Team": "team",
    "AB": "AB",
    "PA": "PA",
    "H": "H",
    "R": "R",
    "HR": "HR",
    "RBI": "RBI",
    "SB": "SB",
    "BB": "BB",
    "HBP": "HBP",
    "AVG": "AVG",
}

PITCHER_COLUMNS = {
    "Name": "name",
    "Team": "team",
    "Starter?": "is_starter",
    "IP": "IP",
    "G": "G",
    "GS": "GS",
    "W": "W",
    "SV": "SV",
    "ERA": "ERA",
}


def load_projections(filepath: Path = DEFAULT_FILE) -> list[Player]:
    """
    Read the PitcherList Excel file and return a list of Player objects
    with projected stats populated.
    """
    if not filepath.exists():
        print(f"[pitcherlist] File not found: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    players: list[Player] = []

    # --- Hitters ---
    if "Hitters" in wb.sheetnames:
        players.extend(_load_sheet(wb["Hitters"], HITTER_COLUMNS, "hitter"))
    else:
        print("[pitcherlist] WARNING: No 'Hitters' sheet found in Excel file.")

    # --- Pitchers ---
    if "Pitchers" in wb.sheetnames:
        players.extend(_load_sheet(wb["Pitchers"], PITCHER_COLUMNS, "pitcher"))
    else:
        print("[pitcherlist] WARNING: No 'Pitchers' sheet found in Excel file.")

    wb.close()
    print(f"[pitcherlist] Loaded {len(players)} players from {filepath.name}")
    return players


def _load_sheet(ws, column_map: dict, sheet_type: str) -> list[Player]:
    """Parse a worksheet into Player objects."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    # Build index: Excel column index -> our internal key
    col_idx = {}
    for i, col_name in enumerate(header):
        if col_name in column_map:
            col_idx[i] = column_map[col_name]

    players = []
    for row in rows[1:]:
        parsed = {}
        for i, key in col_idx.items():
            val = row[i] if i < len(row) else None
            parsed[key] = val

        player = _build_player(parsed, sheet_type)
        if player:
            players.append(player)

    return players


def _build_player(parsed: dict, sheet_type: str) -> Optional[Player]:
    """Convert a parsed row dict into a Player object."""
    name = parsed.get("name")
    team = parsed.get("team")
    if not name or not team:
        return None

    # Determine player_type
    if sheet_type == "hitter":
        player_type = "hitter"
        positions = ["DH"]  # position eligibility comes from players_master.csv
        stat_keys = ["AB", "PA", "H", "R", "HR", "RBI", "SB", "BB", "HBP", "AVG"]
    else:
        is_starter = parsed.get("is_starter")
        player_type = "sp" if is_starter == 1.0 else "rp"
        positions = ["SP"] if player_type == "sp" else ["RP"]
        stat_keys = ["IP", "G", "GS", "W", "SV", "ERA"]

    # Build projected_stats with only numeric values
    projected_stats = {}
    for key in stat_keys:
        val = parsed.get(key)
        if val is not None:
            try:
                projected_stats[key] = float(val)
            except (ValueError, TypeError):
                pass

    return Player(
        name=str(name).strip(),
        team=str(team).strip(),
        positions=positions,
        player_type=player_type,
        projected_stats=projected_stats,
    )


if __name__ == "__main__":
    players = load_projections()
    hitters = [p for p in players if p.player_type == "hitter"]
    sps = [p for p in players if p.player_type == "sp"]
    rps = [p for p in players if p.player_type == "rp"]
    print(f"\nHitters: {len(hitters)}, SP: {len(sps)}, RP: {len(rps)}")

    print("\nTop 10 hitters by projected points:")
    for p in sorted(hitters, key=lambda p: p.projected_points, reverse=True)[:10]:
        print(f"  {p.name:<25} {p.team:<5} {p.projected_points:7.1f} pts  {p.stat_summary()}")

    print("\nTop 10 SP by projected points:")
    for p in sorted(sps, key=lambda p: p.projected_points, reverse=True)[:10]:
        print(f"  {p.name:<25} {p.team:<5} {p.projected_points:7.1f} pts  {p.stat_summary()}")

    print("\nTop 10 RP by projected points:")
    for p in sorted(rps, key=lambda p: p.projected_points, reverse=True)[:10]:
        print(f"  {p.name:<25} {p.team:<5} {p.projected_points:7.1f} pts  {p.stat_summary()}")
